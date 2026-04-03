from flask import Flask, jsonify, request, send_from_directory, abort
from flask_cors import CORS
from datetime import datetime
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / 'enquiries.json'
EXCEL_FILE = BASE_DIR / 'enquiries.xlsx'
HOME_PAGE = 'Sri Vishnu Catering.html'


def save_to_excel(record: dict) -> None:
    headers = ['name', 'email', 'phone', 'eventType', 'message', 'createdAt']

    if EXCEL_FILE.exists():
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Enquiries'
        sheet.append(headers)

    sheet.append([record[key] for key in headers])
    workbook.save(EXCEL_FILE)


@app.get('/')
def serve_home():
    return send_from_directory(BASE_DIR, HOME_PAGE)


@app.get('/<path:page>')
def serve_html_page(page: str):
    # Prevent path traversal and only allow local HTML pages.
    safe_name = Path(page).name
    if safe_name != page or not safe_name.endswith('.html'):
        abort(404)

    target = BASE_DIR / safe_name
    if not target.exists() or not target.is_file():
        abort(404)

    return send_from_directory(BASE_DIR, safe_name)


@app.get('/api/health')
def health() -> tuple:
    return jsonify({'status': 'ok'}), 200


@app.post('/api/book')
def book() -> tuple:
    data = request.get_json(silent=True) or {}

    required = ['name', 'email', 'phone', 'eventType', 'message']
    missing = [field for field in required if not str(data.get(field, '')).strip()]
    if missing:
        return jsonify({'error': f"Missing fields: {', '.join(missing)}"}), 400

    record = {
        'name': data['name'].strip(),
        'email': data['email'].strip(),
        'phone': data['phone'].strip(),
        'eventType': data['eventType'].strip(),
        'message': data['message'].strip(),
        'createdAt': datetime.utcnow().isoformat() + 'Z'
    }

    existing = []
    if DATA_FILE.exists():
        try:
            existing = json.loads(DATA_FILE.read_text(encoding='utf-8'))
            if not isinstance(existing, list):
                existing = []
        except json.JSONDecodeError:
            existing = []

    existing.append(record)
    DATA_FILE.write_text(json.dumps(existing, indent=2), encoding='utf-8')
    save_to_excel(record)

    return jsonify({'message': 'Enquiry received successfully'}), 201


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
